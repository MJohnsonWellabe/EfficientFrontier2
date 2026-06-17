#!/usr/bin/env python3
"""
HISTORICAL / SUPERSEDED (2026-06-14): this was the one-time programmatic migration of the
original `EffFrontierEngine_5.xlsx`. The canonical workbook is now the hand-refined, validated
`EffFrontierEngine_V2Slim_Final.xlsx` (cleaner 2-D product×year Scalars layout); `_5` has been retired.
Kept only as a reference record of the original migration — do NOT run it against V2Slim.

Restructure the EfficientFrontier workbook's `Scalars` sheet into the online
engine's systematic + process decomposition (per experience year, 2026-2055) for
claims, lapse/term and PN NIER, and rewire the recalc tabs to pick them up.

Sales stays as-is (per-year, per-product, keyed on issue year).

Layout written to `Scalars` (sales block rows 1-14 untouched):
  Systematic (single value, col C; identity-default):
    C17/C18/C19  claims systematic  MS/PN/HI            (multiplier, default 1)
    C22/C23/C24  term  systematic   MS/PN/HI            (multiplier, default 1; C23==C18)
    C26          NIER  systematic   PN                  (additive bps, default 0)
  Process (per experience year C:AF; identity-default):
    C29:AF29 / 30 / 31   claims process MS/PN/HI        (multiplier, default 1)
    C34:AF34 / 35 / 36   term   process MS/PN/HI        (multiplier, default 1; row35==row30)
    C39:AF39             NIER   process PN              (additive bps, default 0)

Combination performed by the rewired recalc formulas:
    claims_eff[y] = C17 * claims_proc[y]      (and 18/19)
    term_eff[y]   = C22 * term_proc[y]        (and 23/24)
    NIER new-business (VNB Recalc) = C26 + nier_proc[y]
    NIER back-book   (EV Recalc)   =        nier_proc[y]

With identity defaults the INDEX/MATCH returns the identity, so the recalc
reproduces the baseline byte-for-byte (flat-equivalence invariant).

Idempotent: aborts if the sheet already carries the restructure marker.
Usage: python3 tools/per_year_scalars.py [path-to-xlsx]
"""
import re, sys
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

PATH = sys.argv[1] if len(sys.argv) > 1 else 'EffFrontierEngine_5.xlsx'

YEARS_RANGE = 'Scalars!$C$1:$AF$1'          # experience-year header 2026..2055
LAST_COL = 'AF'                              # 2055

# claims/term token -> process row used in its INDEX lookup
CLAIMS = {'17': 29, '18': 30, '19': 31}     # MS/PN/HI claims systematic -> process row
TERM   = {'22': 34, '23': 35, '24': 36}     # MS/PN/HI term   systematic -> process row

MARKER = 'Claims systematic (multiplier)'   # written to Scalars!B16

def claims_term_repl(colletter, sysnum, procrow):
    """ Scalars!$C$<sysnum>  ->  (sys * IFERROR(INDEX(procrow, MATCH(<col>$1, years)),1)) """
    return ("(Scalars!$C$%s*IFERROR(INDEX(Scalars!$C$%d:$%s$%d,1,"
            "MATCH(%s$1,%s,0)),1))" % (sysnum, procrow, LAST_COL, procrow, colletter, YEARS_RANGE))

def nier_repl(colletter, with_systematic):
    """ Scalars!$C$26 -> (C26 + proc[y])  for new business, or proc[y] only for back book.
        NII column-year is YEAR(<col>3). """
    proc = "IFERROR(INDEX(Scalars!$C$39:$%s$39,1,MATCH(YEAR(%s3),%s,0)),0)" % (LAST_COL, colletter, YEARS_RANGE)
    return ("(Scalars!$C$26+%s)" % proc) if with_systematic else proc

# regex tokens (allow optional $), guard against trailing digits (so $C$1 / $C$12 never match $C$17)
TOK = re.compile(r'Scalars!\$?C\$?(17|18|19|22|23|24|26)(?![0-9])')

def main():
    wb = openpyxl.load_workbook(PATH, data_only=False)
    sc = wb['Scalars']

    if sc['B16'].value == MARKER:
        print('ABORT: workbook already restructured (Scalars!B16 == marker). Nothing to do.')
        return

    # ---- safety: confirm the scalar tokens live only in the three recalc tabs ----
    targets = {'Input EV Recalc', 'VNB Recalc', 'EV Recalc'}
    stray = {}
    for name in wb.sheetnames:
        if name in targets or name == 'Scalars':
            continue
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and TOK.search(cell.value):
                    stray.setdefault(name, cell.coordinate)
    if stray:
        print('ABORT: scalar tokens referenced outside the recalc tabs:', stray)
        return

    # ---- A. restructure Scalars ----
    # extend year header to AF (continue =prev+1)
    for ci in range(column_index_from_string('M'), column_index_from_string(LAST_COL) + 1):
        prev = get_column_letter(ci - 1)
        sc.cell(1, ci).value = '=%s1+1' % prev

    # systematic block (keep cells, relabel)
    sc['B16'] = MARKER
    sc['B17'] = 'Medicare Supplement'; sc['C17'] = 1
    sc['B18'] = 'Preneed (mortality)'; sc['C18'] = 1
    sc['B19'] = 'Hospital Indemnity';  sc['C19'] = 1
    sc['B21'] = 'Term/lapse systematic (multiplier)'
    sc['B22'] = 'Medicare Supplement'; sc['C22'] = 1
    sc['B23'] = 'Preneed (= claims systematic)'; sc['C23'] = '=C18'
    sc['B24'] = 'Hospital Indemnity';  sc['C24'] = 1
    sc['B26'] = 'NIER systematic — PN only (additive bps; e.g. -0.0035 = -35bps)'; sc['C26'] = 0

    # clear any old notes rows
    for r in range(27, 45):
        for ci in range(2, column_index_from_string(LAST_COL) + 1):
            sc.cell(r, ci).value = None

    def fill(row, label, default, formula_link=None):
        sc.cell(row, 2).value = label
        for ci in range(column_index_from_string('C'), column_index_from_string(LAST_COL) + 1):
            if formula_link:
                src = get_column_letter(ci)
                sc.cell(row, ci).value = '=%s%d' % (src, formula_link)
            else:
                sc.cell(row, ci).value = default

    sc['B28'] = 'Claims process (multiplier, by experience year 2026-2055)'
    fill(29, 'Medicare Supplement', 1)
    fill(30, 'Preneed (mortality)', 1)
    fill(31, 'Hospital Indemnity', 1)
    sc['B33'] = 'Term/lapse process (multiplier, by experience year)'
    fill(34, 'Medicare Supplement', 1)
    fill(35, 'Preneed (= claims process)', None, formula_link=30)   # PN term proc == PN claims proc
    fill(36, 'Hospital Indemnity', 1)
    sc['B38'] = 'NIER process — PN only (additive bps, by experience year)'
    fill(39, 'Preneed', 0)
    sc['B41'] = ('NOTE: claims/term effective = systematic x process[year]; '
                 'NIER new business (VNB) = systematic + process[year]; NIER back book (EV) = process[year].')

    # ---- B. rewire the recalc tabs ----
    counts = {}
    def rewire(sheet, kind):
        ws = wb[sheet]; n = 0
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or 'Scalars!' not in v:
                    continue
                if not TOK.search(v):
                    continue
                col = cell.column_letter
                def sub(m):
                    num = m.group(1)
                    if num in CLAIMS:
                        return claims_term_repl(col, num, CLAIMS[num])
                    if num in TERM:
                        return claims_term_repl(col, num, TERM[num])
                    if num == '26':
                        return nier_repl(col, with_systematic=(kind == 'nier_nb'))
                    return m.group(0)
                new = TOK.sub(sub, v)
                if new != v:
                    cell.value = new; n += 1
        counts[sheet] = n

    rewire('Input EV Recalc', 'cl')      # claims (17/18/19) + lapse (22/23/24)
    rewire('VNB Recalc', 'nier_nb')      # NIER new business: systematic + process
    rewire('EV Recalc', 'nier_bb')       # NIER back book: process only

    wb.save(PATH)
    print('Saved', PATH)
    print('Rewired cells:', counts)
    # echo a couple of samples
    wb2 = openpyxl.load_workbook(PATH, data_only=False)
    print('SAMPLE Input EV Recalc!E8 :', wb2['Input EV Recalc']['E8'].value)
    print('SAMPLE Input EV Recalc!R9 :', wb2['Input EV Recalc']['R9'].value)
    print('SAMPLE VNB Recalc!AI5      :', wb2['VNB Recalc']['AI5'].value)
    print('SAMPLE EV Recalc!AJ5       :', wb2['EV Recalc']['AJ5'].value)

if __name__ == '__main__':
    main()
